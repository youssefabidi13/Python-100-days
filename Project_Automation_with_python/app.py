import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb=xl.load_workbook('transactions.xlsx')
sheet =wb['Sheet1']
cell = sheet['a1']
print(cell.value)
for row in range(2,sheet.max_row +1):
    cell = sheet.cell(row,3)
    print(cell.value)
    corrected_price = cell.value * 0.9
    print(corrected_price)
    corrected_price_cell = sheet.cell(row,4)
    corrected_price_cell.value = corrected_price

Values = Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,min_col=4,
          max_col=4)

chart = BarChart()
chart.add_data(Values)
sheet.add_chart(chart,'e2')
wb.save('transactions2.xlsx')
